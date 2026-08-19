#!/usr/bin/env python3
"""회신된 전문가 검토지를 폴더째 읽어 문항별로 집계한다.

한 파일이 검토자 한 명이다. 파일명은 제각각이어도 된다 — 폴더 안의 .xlsx/.csv를 전부 읽는다.

내는 값
  I-CVI   문항 하나를 「이 유형이 맞다」고 본 검토자의 비율
  기준    0.78 미만 표시(요청받은 선)과, 검토자 수에 맞는 임계값(Lynn 1986) 두 가지를 함께 본다
  의견    자유 의견을 문항별로 모아 누가 무슨 말을 했는지 붙여 둔다

쓰는 법
  python scripts/aggregate_review.py <회신폴더> [--out 집계.csv] [--comment-col H]
"""
from __future__ import annotations

import argparse
import csv
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# 요청받은 표시선. 이 값만으로 판정하지 않는다 — 아래 LYNN_MINIMUM을 함께 본다.
REQUESTED_THRESHOLD = 0.78

# 검토자 수별 I-CVI 임계값(Lynn 1986). 우연 일치를 배제하려면 사람이 적을수록 만장일치에 가까워야 한다.
# 5명 이하에서는 전원이 「맞다」라고 해야 한다. 0.78은 9명 이상일 때의 선이다.
LYNN_MINIMUM = {2: 1.00, 3: 1.00, 4: 1.00, 5: 1.00, 6: 0.83, 7: 0.86, 8: 0.88, 9: 0.78, 10: 0.80}


def lynn_minimum(raters: int) -> float:
    """검토자 수에 맞는 임계값. 표에 없는 큰 수는 9명 이상 기준인 0.78을 쓴다."""
    if raters in LYNN_MINIMUM:
        return LYNN_MINIMUM[raters]
    return 0.78 if raters > 10 else 1.00


def norm(value) -> str:
    """헤더·응답 비교용. 공백과 전각/반각 차이로 어긋나지 않게 맞춘다."""
    if value is None:
        return ''
    text = unicodedata.normalize('NFKC', str(value)).strip()
    return ' '.join(text.split())


# 「맞다」로 볼 응답과 「아니다」로 볼 응답. 숫자는 척도를 보고 따로 판단한다(read_verdict 참고).
AGREE_WORDS = {'예', '네', '맞다', '맞음', '적합', '해당', '동의', 'o', '○', '◯', 'y', 'yes', 'true', 'v', '✓'}
DENY_WORDS = {'아니오', '아니요', '아님', '틀림', '틀리다', '부적합', '비해당', '비동의', 'x', '✕', '×', 'n', 'no', 'false'}

# 헤더를 이름으로 찾는다. 자리(열 순서)가 조금 달라도 읽히게 하기 위해서다.
ID_HINTS = ('번호', '문항번호', '문항 번호', 'id', 'no')
ITEM_HINTS = ('문항', '내용', '항목', '질문')
VERDICT_HINTS = ('유형', '적합', '관련', '판정', '타당', 'cvi')
COMMENT_HINTS = ('의견', '코멘트', '비고', '메모', '자유')


def find_column(headers: list[str], hints: tuple[str, ...]) -> int | None:
    """헤더 이름으로 열을 찾는다. 없으면 None."""
    lowered = [norm(h).lower() for h in headers]
    for index, head in enumerate(lowered):
        if head and any(hint in head for hint in hints):
            return index
    return None


def read_rows(path: Path) -> list[list]:
    """한 벌을 2차원 리스트로 읽는다. .xlsx는 첫 시트만 본다."""
    if path.suffix.lower() == '.csv':
        for encoding in ('utf-8-sig', 'cp949'):
            try:
                with path.open(encoding=encoding, newline='') as handle:
                    return [row for row in csv.reader(handle)]
            except UnicodeDecodeError:
                continue
        raise SystemExit(f'{path.name}: 글자코드를 읽지 못했습니다(utf-8·cp949 모두 실패).')
    if load_workbook is None:
        raise SystemExit('xlsx를 읽으려면 openpyxl이 필요합니다: pip install openpyxl')
    book = load_workbook(path, data_only=True, read_only=True)
    try:
        return [list(row) for row in book[book.sheetnames[0]].iter_rows(values_only=True)]
    finally:
        book.close()


def header_row_index(rows: list[list]) -> int:
    """머리글 줄을 찾는다. 검토지 위쪽에 안내 문구가 몇 줄 있어도 넘어간다."""
    for index, row in enumerate(rows[:20]):
        cells = [norm(cell).lower() for cell in row]
        if any(cell and any(hint in cell for hint in VERDICT_HINTS) for cell in cells):
            return index
    return 0


class Sheet:
    """회신 한 벌. 문항 번호 → (판정, 의견)."""

    def __init__(self, path: Path, comment_col: int | None):
        self.path = path
        self.name = path.stem
        rows = read_rows(path)
        if not rows:
            raise SystemExit(f'{path.name}: 내용이 비어 있습니다.')
        head = header_row_index(rows)
        headers = [norm(cell) for cell in rows[head]]
        id_col = find_column(headers, ID_HINTS)
        item_col = find_column(headers, ITEM_HINTS)
        verdict_col = find_column(headers, VERDICT_HINTS)
        if verdict_col is None:
            raise SystemExit(f'{path.name}: 「이 유형이 맞다」 판정 열을 찾지 못했습니다. '
                             f'머리글에 {" · ".join(VERDICT_HINTS)} 중 하나가 있어야 합니다.')
        # 자유 의견은 요청대로 H열을 먼저 본다. 지정이 없으면 머리글 이름으로 찾는다.
        if comment_col is None:
            comment_col = find_column(headers, COMMENT_HINTS)
        self.headers = headers
        self.verdict_header = headers[verdict_col] if verdict_col < len(headers) else ''
        self.entries: dict[str, dict] = {}
        self.order: list[str] = []
        for row in rows[head + 1:]:
            def cell(index):
                return row[index] if index is not None and index < len(row) else None
            key = norm(cell(id_col)) if id_col is not None else ''
            text = norm(cell(item_col))
            if not key:
                key = text
            if not key:
                continue  # 번호도 내용도 없는 줄은 빈 줄이다.
            raw = cell(verdict_col)
            if norm(raw) == '' and text == '':
                continue
            self.entries[key] = {'item': text, 'raw': raw, 'comment': norm(cell(comment_col))}
            self.order.append(key)


def detect_scale(values: list) -> str:
    """숫자 응답이 2점(0/1)인지 4점 척도인지 정한다.

    같은 「1」이 척도에 따라 정반대를 뜻한다 — 2점에서는 「맞다」, 4점에서는 「전혀 아니다」다.
    그래서 한 칸만 보고 정하지 않고 열 전체의 값 집합을 보고 정한다.
    """
    numbers = set()
    for value in values:
        text = norm(value)
        if text == '':
            continue
        try:
            numbers.add(float(text))
        except ValueError:
            return 'words'
    if not numbers:
        return 'words'
    if numbers <= {0.0, 1.0}:
        return 'binary'
    if numbers <= {1.0, 2.0, 3.0, 4.0}:
        return 'likert4'
    if numbers <= {1.0, 2.0, 3.0, 4.0, 5.0}:
        return 'likert5'
    return 'unknown'


def read_verdict(raw, scale: str):
    """한 칸을 True(맞다)/False(아니다)/None(미응답)으로 바꾼다."""
    text = norm(raw)
    if text == '':
        return None
    lowered = text.lower()
    if lowered in AGREE_WORDS:
        return True
    if lowered in DENY_WORDS:
        return False
    try:
        number = float(text)
    except ValueError:
        return None
    if scale == 'binary':
        return number == 1.0
    if scale == 'likert4':
        return number >= 3.0  # 4점 척도는 3·4를 「관련 있다」로 본다.
    if scale == 'likert5':
        return number >= 4.0
    return None


def aggregate(folder: Path, comment_col: int | None):
    paths = sorted(p for p in folder.iterdir()
                   if p.suffix.lower() in {'.xlsx', '.csv'} and not p.name.startswith('~$'))
    if not paths:
        raise SystemExit(f'{folder}: 읽을 회신 파일(.xlsx·.csv)이 없습니다.')
    sheets = [Sheet(path, comment_col) for path in paths]

    # 문항 목록은 첫 벌의 순서를 따른다. 다른 벌에만 있는 문항은 뒤에 붙이고 따로 알린다.
    keys: list[str] = list(sheets[0].order)
    extra: dict[str, list[str]] = {}
    for sheet in sheets[1:]:
        for key in sheet.order:
            if key not in keys:
                keys.append(key)
                extra.setdefault(key, []).append(sheet.name)
    missing: dict[str, list[str]] = {}
    for key in keys:
        absent = [sheet.name for sheet in sheets if key not in sheet.entries]
        if absent:
            missing[key] = absent

    # 척도는 전체 판정 칸을 모아 한 번만 정한다. 벌마다 다르게 읽히면 집계가 어긋난다.
    scale = detect_scale([entry['raw'] for sheet in sheets for entry in sheet.entries.values()])

    items = []
    for key in keys:
        text = ''
        agree = deny = blank = 0
        comments = []
        for sheet in sheets:
            entry = sheet.entries.get(key)
            if entry is None:
                blank += 1
                continue
            text = text or entry['item']
            verdict = read_verdict(entry['raw'], scale)
            if verdict is True:
                agree += 1
            elif verdict is False:
                deny += 1
            else:
                blank += 1
            if entry['comment']:
                comments.append((sheet.name, entry['comment'], verdict))
        answered = agree + deny
        icvi = (agree / answered) if answered else None
        items.append({'key': key, 'item': text, 'agree': agree, 'deny': deny, 'blank': blank,
                      'answered': answered, 'icvi': icvi, 'comments': comments})
    return sheets, scale, items, missing, extra


def report(sheets, scale, items, missing, extra, threshold: float):
    raters = len(sheets)
    strict = lynn_minimum(raters)
    scale_label = {'binary': '2점(0·1)', 'likert4': '4점 척도(3·4를 「맞다」로)',
                   'likert5': '5점 척도(4·5를 「맞다」로)', 'words': '예·아니오 문구',
                   'unknown': '알 수 없음'}.get(scale, scale)

    print(f'회신 {raters}벌 · 문항 {len(items)}개 · 판정 {scale_label}')
    print('회신 파일: ' + ' · '.join(sheet.name for sheet in sheets))
    print(f'판정 열 이름: {sheets[0].verdict_header or "(이름 없음)"}')
    print()
    print(f'표시선 {threshold:.2f} (요청받은 값) · 검토자 {raters}명에 맞는 임계값 {strict:.2f} (Lynn 1986)')
    if strict > threshold:
        print(f'  ※ 검토자가 {raters}명이면 {strict:.2f} 미만은 통과로 볼 수 없다. '
              f'{threshold:.2f}만 보면 실제로는 미달인 문항을 통과로 셀 수 있다.')
    print()

    header = f'{"번호":<8}{"맞다":>6}{"아니다":>7}{"무응답":>7}{"I-CVI":>8}  판정'
    print(header)
    print('-' * (len(header) + 12))
    for row in items:
        icvi = row['icvi']
        shown = f'{icvi:.3f}' if icvi is not None else '  —  '
        if icvi is None:
            mark = '집계 불가(응답 없음)'
        elif icvi < threshold:
            mark = f'★ {threshold:.2f} 미만'
        elif icvi < strict:
            mark = f'△ {threshold:.2f}는 넘으나 {strict:.2f} 미달'
        else:
            mark = '통과'
        # 일부만 답한 문항은 비율이 높아도 통과로 읽으면 안 된다. 분모가 다른 값이다.
        if row['answered'] and row['answered'] < raters:
            mark += f'  ※ {raters}명 중 {row["answered"]}명만 답함 — 다른 문항과 같은 선으로 보지 말 것'
        print(f'{row["key"]:<8}{row["agree"]:>6}{row["deny"]:>7}{row["blank"]:>7}{shown:>8}  {mark}')

    partial = [row for row in items if row['answered'] and row['answered'] < raters]
    # 전원이 답한 문항만 S-CVI에 넣는다. 분모가 다른 값을 한 평균에 섞으면 숫자가 실제보다 좋아 보인다.
    scored = [row['icvi'] for row in items if row['icvi'] is not None and row['answered'] == raters]
    if scored:
        print()
        print(f'S-CVI/Ave {sum(scored) / len(scored):.3f} · '
              f'S-CVI/UA {sum(1 for v in scored if v >= 1.0) / len(scored):.3f} (전원 일치 문항 비율)'
              + (f' · {len(scored)}개 문항으로 계산' if partial else ''))
        if partial:
            print(f'  ※ 일부만 답한 문항 {len(partial)}개는 뺐다: '
                  + ' · '.join(f'{r["key"]}({r["answered"]}/{raters})' for r in partial))
        low = [row for row in items if row['icvi'] is not None and row['icvi'] < threshold]
        gap = [row for row in items if row['icvi'] is not None and threshold <= row['icvi'] < strict]
        print(f'{threshold:.2f} 미만 {len(low)}개' + (': ' + ' · '.join(r['key'] for r in low) if low else ''))
        if gap:
            print(f'{threshold:.2f}~{strict:.2f} 사이 {len(gap)}개(검토자 수 기준으로는 미달): '
                  + ' · '.join(r['key'] for r in gap))

    if missing or extra:
        print()
        print('맞물리지 않은 문항')
        for key, names in missing.items():
            if key in extra:
                continue  # 아래에서 「~에만 있음」으로 한 번만 적는다.
            print(f'  {key}: 빠진 회신 {" · ".join(names)}')
        for key, names in extra.items():
            print(f'  {key}: {" · ".join(names)}에만 있음')
        print('  ※ 줄 순서로 억지로 짝짓지 않았다. 번호가 다르면 다른 문항으로 센다.')

    print()
    print('자유 의견')
    any_comment = False
    for row in items:
        if not row['comments']:
            continue
        any_comment = True
        print(f'  [{row["key"]}] {row["item"][:40]}')
        for name, comment, verdict in row['comments']:
            tag = {True: '맞다', False: '아니다', None: '무응답'}[verdict]
            print(f'    - {name} ({tag}): {comment}')
    if not any_comment:
        print('  (없음)')


def write_csv(path: Path, items, threshold: float, strict: float):
    with path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.writer(handle)
        writer.writerow(['문항번호', '문항', '맞다', '아니다', '무응답', 'I-CVI',
                         f'{threshold:.2f}미만', f'{strict:.2f}미달', '자유의견'])
        for row in items:
            icvi = row['icvi']
            writer.writerow([
                row['key'], row['item'], row['agree'], row['deny'], row['blank'],
                '' if icvi is None else f'{icvi:.3f}',
                'Y' if icvi is not None and icvi < threshold else '',
                'Y' if icvi is not None and icvi < strict else '',
                ' / '.join(f'{name}: {comment}' for name, comment, _ in row['comments'])
            ])


def main():
    parser = argparse.ArgumentParser(description='회신 검토지를 폴더째 읽어 문항별 I-CVI를 낸다.')
    parser.add_argument('folder', type=Path, help='회신 파일이 든 폴더')
    parser.add_argument('--out', type=Path, default=None, help='집계 결과를 CSV로도 저장할 경로')
    parser.add_argument('--threshold', type=float, default=REQUESTED_THRESHOLD, help='표시선(기본 0.78)')
    parser.add_argument('--comment-col', default=None,
                        help='자유 의견 열을 엑셀 열문자로 지정(예: H). 없으면 머리글 이름으로 찾는다.')
    args = parser.parse_args()

    if not args.folder.is_dir():
        raise SystemExit(f'{args.folder}: 폴더가 아닙니다.')
    comment_col = None
    if args.comment_col:
        letter = args.comment_col.strip().upper()
        if not letter.isalpha():
            raise SystemExit('--comment-col 은 엑셀 열문자여야 합니다(예: H).')
        comment_col = 0
        for char in letter:
            comment_col = comment_col * 26 + (ord(char) - ord('A') + 1)
        comment_col -= 1

    sheets, scale, items, missing, extra = aggregate(args.folder, comment_col)
    report(sheets, scale, items, missing, extra, args.threshold)
    if args.out:
        write_csv(args.out, items, args.threshold, lynn_minimum(len(sheets)))
        print()
        print(f'CSV 저장: {args.out}')


if __name__ == '__main__':
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')  # 윈도우 콘솔이 cp949라 한글이 깨진다.
    main()
